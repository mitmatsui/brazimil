import json
import openpyxl
from datetime import datetime

print("Lendo o Excel...")

wb = openpyxl.load_workbook("SIMULADOR TINTAS MIL BRAZIMIL 1.xlsx", data_only=True)

ws = wb["Custo"]
dados_custo = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    nome  = row[0]
    custo = row[1]
    if nome and custo and isinstance(custo, (int, float)) and custo > 0:
        dados_custo[nome] = round(custo, 4)

ws2 = wb["Peso"]
dados_peso = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    nome = row[0]
    peso = row[1]
    if nome and peso and isinstance(peso, (int, float)) and peso > 0:
        dados_peso[nome] = round(peso, 4)

print(f"Custos lidos: {len(dados_custo)} produtos")
print(f"Pesos lidos:  {len(dados_peso)} produtos")

tabela = {
    "BRAZIMIL ECONÔMICA 3,0L REFIL BRANCO NEVE":      ("Tinta Base Água", "Econômica",        9.50),
    "BRAZIMIL ECONÔMICA 3,0L REFIL COR":               ("Tinta Base Água", "Econômica",        9.50),
    "BRAZIMIL ECONÔMICA 3,0L GALÃO BRANCO NEVE":       ("Tinta Base Água", "Econômica",       17.99),
    "BRAZIMIL ECONÔMICA 3,0L GALÃO COR":               ("Tinta Base Água", "Econômica",       17.99),
    "BRAZIMIL ECONÔMICA 15L BALDE BRANCO NEVE":        ("Tinta Base Água", "Econômica",       58.00),
    "BRAZIMIL ECONÔMICA 15L BALDE COR":                ("Tinta Base Água", "Econômica",       69.90),
    "BRAZIMIL EXTERNA 3,0L BALDE BRANCO NEVE":         ("Tinta Base Água", "Externa",         26.00),
    "BRAZIMIL EXTERNA 3,0L BALDE COR":                 ("Tinta Base Água", "Externa",         24.20),
    "BRAZIMIL EXTERNA 15L BALDE BRANCO NEVE":          ("Tinta Base Água", "Externa",         94.00),
    "BRAZIMIL EXTERNA 15L BALDE COR":                  ("Tinta Base Água", "Externa",         90.11),
    "BRAZIMIL SEMI BRILHO 3,0L BALDE BRANCO NEVE":     ("Tinta Base Água", "Semi Brilho",     43.00),
    "BRAZIMIL SEMI BRILHO 3,0L BALDE COR":             ("Tinta Base Água", "Semi Brilho",     35.00),
    "BRAZIMIL SEMI BRILHO 15L BALDE BRANCO NEVE":      ("Tinta Base Água", "Semi Brilho",    170.00),
    "BRAZIMIL SEMI BRILHO 15L BALDE COR":              ("Tinta Base Água", "Semi Brilho",    170.00),
    "BRAZIMIL TINTA PISO 3,0L BALDE BRANCO NEVE":      ("Tinta Base Água", "Piso",            30.00),
    "BRAZIMIL TINTA PISO 3,0L BALDE COR":              ("Tinta Base Água", "Piso",            28.50),
    "BRAZIMIL TINTA PISO 15L BALDE BRANCO NEVE":       ("Tinta Base Água", "Piso",           130.00),
    "BRAZIMIL TINTA PISO 15L BALDE COR":               ("Tinta Base Água", "Piso",           108.00),
    "BRAZIMIL SELADOR ACRIL 3,0L BALDE PIGMENTADO":    ("Tinta Base Água", "Selador Acril",   17.99),
    "BRAZIMIL SELADOR ACRIL 15L BALDE PIGMENTADO":     ("Tinta Base Água", "Selador Acril",   51.00),
    "BRAZIMIL TINTA P GESSO 3,0L":                     ("Tinta Base Água", "Gesso",           20.00),
    "BRAZIMIL TINTA P GESSO 15L":                      ("Tinta Base Água", "Gesso",           73.00),
    "LIQUIBRILHO 3,0L BALDE INCOLOR":                  ("Tinta Base Água", "Liquibrilho",     32.00),
    "LIQUIBRILHO 15L BALDE INCOLOR":                   ("Tinta Base Água", "Liquibrilho",    135.00),
    "BRAZIMIL ESMALTE SINTÉTICO 0,75L LATA ALUMÍNIO":  ("Tinta Base Solvente", "Esmalte Sint.", 27.20),
    "BRAZIMIL ESMALTE SINTÉTICO 0,75L LATA BRANCO":    ("Tinta Base Solvente", "Esmalte Sint.", 16.11),
    "BRAZIMIL ESMALTE SINTÉTICO 0,75L LATA PRETO":     ("Tinta Base Solvente", "Esmalte Sint.", 16.87),
    "BRAZIMIL ESMALTE SINTÉTICO 0,75L LATA COR":       ("Tinta Base Solvente", "Esmalte Sint.", 16.87),
    "BRAZIMIL ESMALTE SINTÉTICO 3,0L LATA ALUMÍNIO":   ("Tinta Base Solvente", "Esmalte Sint.", 99.00),
    "BRAZIMIL ESMALTE SINTÉTICO 3,0L LATA BRANCO":     ("Tinta Base Solvente", "Esmalte Sint.", 52.26),
    "BRAZIMIL ESMALTE SINTÉTICO 3,0L LATA PRETO":      ("Tinta Base Solvente", "Esmalte Sint.", 39.90),
    "BRAZIMIL ESMALTE SINTÉTICO 3,0L LATA COR":        ("Tinta Base Solvente", "Esmalte Sint.", 46.90),
    "BRAZIMIL ESMALTE SINTÉTICO 3,0L LATA FOSCO":      ("Tinta Base Solvente", "Esmalte Sint.", 54.50),
    "ESMALTE BASE ÁGUA 0,75L":                         ("Tinta Base Solvente", "Esmalte Base Água", 20.00),
    "ESMALTE BASE ÁGUA 3,0L":                          ("Tinta Base Solvente", "Esmalte Base Água", 80.00),
    "BRAZIMIL VERNIZ 0,75L LATA INCOLOR":              ("Tinta Base Solvente", "Verniz",       20.50),
    "BRAZIMIL VERNIZ 0,75L LATA PIGMENTADO":           ("Tinta Base Solvente", "Verniz",       20.50),
    "BRAZIMIL VERNIZ 3,0L LATA INCOLOR":               ("Tinta Base Solvente", "Verniz",       63.00),
    "BRAZIMIL VERNIZ 3,0L LATA PIGMENTADO":            ("Tinta Base Solvente", "Verniz",       63.00),
    "BRAZIMIL ZARCÃO 0,75L LATA BRANCO":               ("Tinta Base Solvente", "Zarcão",       15.90),
    "BRAZIMIL ZARCÃO 0,75L LATA CINZA":                ("Tinta Base Solvente", "Zarcão",       14.90),
    "BRAZIMIL ZARCÃO 0,75L LATA COR":                  ("Tinta Base Solvente", "Zarcão",       14.00),
    "BRAZIMIL ZARCÃO 3,0L LATA CINZA":                 ("Tinta Base Solvente", "Zarcão",       45.55),
    "BRAZIMIL ZARCÃO  18L LATA COR":                   ("Tinta Base Solvente", "Zarcão",      200.00),
    "BRAZIMIL SELADOR P/ MAD. 0,75L LATA INCOLOR":     ("Tinta Base Solvente", "Selador Mad.", 19.00),
    "BRAZIMIL SELADOR P/ MAD. 3,0L LATA INCOLOR":      ("Tinta Base Solvente", "Selador Mad.", 68.00),
    "BRAZIMIL CORRIDA POTE 1KG":                        ("Massas e Texturas", "Corrida",        5.50),
    "BRAZIMIL CORRIDA GALÃO 5KG":                       ("Massas e Texturas", "Corrida",       14.90),
    "BRAZIMIL CORRIDA SACO VALV 10 KG":                 ("Massas e Texturas", "Corrida",       11.00),
    "BRAZIMIL CORRIDA BALDE 20KG":                      ("Massas e Texturas", "Corrida",       36.50),
    "BRAZIMIL ACRILICA POTE 1KG":                       ("Massas e Texturas", "Acrílica",       6.00),
    "BRAZIMIL ACRILICA GALÃO 5KG":                      ("Massas e Texturas", "Acrílica",      19.16),
    "BRAZIMIL ACRILICA SACO VALV 10,0KG":               ("Massas e Texturas", "Acrílica",      27.00),
    "BRAZIMIL ACRILICA BALDE 20KG":                     ("Massas e Texturas", "Acrílica",      62.26),
    "VEDAMIL / MANTA LÍQUIDA BALDE 3,6KG":             ("Complementos", "Vedamil",             32.50),
    "VEDAMIL / MANTA LÍQUIDA BALDE 12KG":              ("Complementos", "Vedamil",            133.00),
    "VEDAMIL / MANTA LÍQUIDA BALDE 15KG":              ("Complementos", "Vedamil",            100.00),
    "VEDAMIL / MANTA LÍQUIDA BALDE 18KG":              ("Complementos", "Vedamil",            150.00),
    "BRAZIMIL BORRACHA LIQUIDA 3,6 KG":                ("Complementos", "Borracha Líquida",    86.67),
    "BRAZIMIL BORRACHA LIQUIDA 14,0 KG":               ("Complementos", "Borracha Líquida",   285.00),
    "CIMENTO QUEIMADO GALÃO":                           ("Complementos", "Cimento Queimado",    40.00),
    "BRAZIMIL COLA BRANCA 0,5KG":                       ("Complementos", "Cola e Solvente",     7.00),
    "BRAZIMIL COLA BRANCA 1,0KG":                       ("Complementos", "Cola e Solvente",    11.50),
    "AGUARRAS 0,900":                                   ("Complementos", "Cola e Solvente",    15.00),
    "AGUARRAS 0,450":                                   ("Complementos", "Cola e Solvente",     7.80),
}

produtos = []
for nome_excel, (cat, sub, preco) in tabela.items():
    custo = dados_custo.get(nome_excel)
    peso  = dados_peso.get(nome_excel)
    if custo and peso:
        nome_exib = nome_excel.replace("BRAZIMIL ", "").strip().title()
        produtos.append({
            "nome":  nome_exib,
            "cat":   cat,
            "sub":   sub,
            "preco": preco,
            "custo": round(custo, 2),
            "peso":  round(peso, 3),
        })

print(f"Produtos no dashboard: {len(produtos)}")

# ── Gera o JSON dos dados ────────────────────────────────────────────────
produtos_json = json.dumps(produtos, ensure_ascii=False)
data_hoje     = datetime.now().strftime("%d/%m/%Y %H:%M")
total         = len(produtos)

# ── Lê o template JS/HTML externo ───────────────────────────────────────
with open("template.html", "r", encoding="utf-8") as f:
    template = f.read()

# ── Substitui apenas os placeholders ────────────────────────────────────
html = template.replace("__PRODUTOS_JSON__", produtos_json)
html = html.replace("__DATA_HOJE__", data_hoje)
html = html.replace("__TOTAL__", str(total))

with open("dashboard_brazimil.html", "w", encoding="utf-8") as f:
    f.write(html)

print("=" * 55)
print("Dashboard v3 gerado: dashboard_brazimil.html")
print("3 abas: Painel Geral · Simulador de Preço · Comparador")
print("=" * 55)
